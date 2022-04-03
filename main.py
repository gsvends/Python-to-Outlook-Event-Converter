# Misc imports
import datetime
from time import strftime

# Outlook imports
from win32com.client import Dispatch
outlook = win32com.client.Dispatch("Outlook.Application")

# Excel Imports
from openpyxl import load_workbook

# This function takes in an excel worksheet and a row to read
# It returns a dictionary containing the values of the row
def getRowValues(ws, row):

    rowDict = {}

    # Load the Rows values into the dictionary
    rowDict['ticketNumberVal'] = ws['A{0}'.format(row)].value
    rowDict['dateVal'] = ws['B{0}'.format(row)].value
    rowDict['mtceTimeCtVal'] = ws['C{0}'.format(row)].value
    rowDict['descriptionVal'] = ws['D{0}'.format(row)].value
    rowDict['circuitIdVal'] = ws['E{0}'.format(row)].value
    rowDict['attSiteIdVal'] = ws['F{0}'.format(row)].value
    rowDict['siteVal'] = ws['G{0}'.format(row)].value
    rowDict['ceRouterVal'] = ws['H{0}'.format(row)].value
    rowDict['ipAddresVal'] = ws['I{0}'.format(row)].value
    rowDict['riskVal'] = ws['J{0}'.format(row)].value
    rowDict['redundancyVal'] = ws['K{0}'.format(row)].value
    rowDict['vTmTicketVal'] = ws['L{0}'.format(row)].value
    rowDict['lcmCommentsVal'] = ws['M{0}'.format(row)].value

    return rowDict

# This function takes a time range and determines the start time and the duration
def parseStartAndDuration(timeString):
    times = timeString.split(' - ')

    startDateTime = datetime.strptime(times[0],'%H:%M%P - %Z')
    endDateTime = datetime.strptime(times[1],'%H:%M%P - %Z')

    rtnDict = {
        startTime : startDateTime,
        duration : (endDateTime - startDateTime).total_seconds() / 60.0
    }

    return rtnDict


# This function takes in a dictionary of excel row data and creates an event in outlook
def createEventFromDictionary(dict):

    # This was for debug just to print the values that were read in.  
    # print('Row Data: {0}'.format(dict))

    # Create the appointment    
    appt = outlook.CreateItem(1) # AppointmentItem

    # Get the start time and duration from the time value    
    evenTimeDict = parseStartAndDuration(dict['mtceTimeCtVal'])
    meetingTimeString = '{0} {1}'.format(dict['dateVal'], evenTimeDict['startTime'])

    appt.Start = meetingTimeString # yyyy-MM-dd hh:mm
    appt.Duration = evenTimeDict['duration'] # In minutes (60 Minutes)

    appt.Subject = dict['descriptionVal'] # TODO: This subject could be anything
    appt.Location = "Zoom" # TODO: What should the location be?
    appt.MeetingStatus = 1 # 1 - olMeeting; Changing the appointment to meeting. Only after changing the meeting status recipients can be added

    # TODO: Add your default Recipients here
    appt.Recipients.Add("test@test.com") # Don't end ; as delimiter

    appt.Save()
    #appt.Send()

# This is the MAIN function that will be called at the end of the file.
def main():
    wb = load_workbook('myFile.xlsx')

    # grab the active worksheet
    ws = wb.active

    # Loop thought the Rows
    prevTicketNumberVal = None
    row = 1
    fileDone = False
    while fileDone == False:

        # Load a row's worth of values
        rowDict = getRowValues(ws, row)

        # See if we've found 2 blank rows in a row... if we have the file is done.
        if rowDict['ticketNumberVal'] == None and prevTicketNumberVal == None:
            fileDone = True
        # See if this row isn't blank... If it isn't, we got data!
        elif rowDict['ticketNumberVal'] != None and rowDict['ticketNumberVal'] != 'Ticket # or Chg #':
            createEventFromDictionary(rowDict)

        row += 1
        prevTicketNumberVal = rowDict['ticketNumberVal']
            
    # Close the file
    wb.close()

# Call the main function to start
main()
